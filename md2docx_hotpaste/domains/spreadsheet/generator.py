"""Spreadsheet file generator - creates XLSX files from table data."""

from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

from ...utils.logging import log
from ...core.errors import InsertError
from .formatting import CellFormat


class SpreadsheetGenerator:
    """表格生成器 - 生成 XLSX 文件（支持复杂格式）"""
    
    @staticmethod
    def generate_xlsx(table_data: List[List[str]], output_path: str, keep_format: bool = True) -> bool:
        """
        从表格数据生成 XLSX 文件（支持 Markdown 格式）
        
        Args:
            table_data: 二维数组表格数据
            output_path: 输出 XLSX 文件路径
            keep_format: 是否保留 Markdown 格式（粗体、斜体等）
            
        Returns:
            True 如果成功生成
        """
        try:
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            if not table_data:
                log("Table data is empty, creating empty spreadsheet")
                wb.save(output_path)
                log(f"Successfully generated XLSX: {output_path}")
                return True
            
            # 设置样式
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            code_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # 写入数据
            for row_idx, row_data in enumerate(table_data, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if keep_format:
                        # 解析 Markdown 格式
                        cell_format = CellFormat(cell_value)
                        clean_text = cell_format.parse()
                        
                        # 检查是否有超链接
                        hyperlink_url = None
                        if cell_format.segments:
                            # 查找第一个超链接
                            for seg in cell_format.segments:
                                if seg.hyperlink_url:
                                    hyperlink_url = seg.hyperlink_url
                                    break
                        
                        # 应用格式
                        if cell_format.has_newline:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        
                        if cell_format.is_code_block:
                            # 代码块样式
                            cell.value = clean_text
                            cell.font = Font(name="Consolas")
                            cell.fill = code_fill
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                        elif hyperlink_url:
                            # 有超链接：添加超链接并设置蓝色下划线样式
                            cell.value = clean_text
                            cell.hyperlink = hyperlink_url
                            cell.font = Font(color="0563C1", underline="single")  # Excel 默认超链接颜色
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        elif len(cell_format.segments) > 1:
                            # 多个片段，使用富文本
                            rich_text_parts = []
                            has_inline_code = False
                            
                            for seg in cell_format.segments:
                                if not seg.text:
                                    continue
                                
                                # 检查是否有行内代码
                                if seg.is_code:
                                    has_inline_code = True
                                
                                # 创建内联字体样式
                                inline_font = InlineFont(
                                    b=seg.bold,
                                    i=seg.italic,
                                    strike=seg.strikethrough,
                                    rFont="Consolas" if seg.is_code else None
                                )
                                
                                # 添加文本块
                                rich_text_parts.append(TextBlock(inline_font, seg.text))
                            
                            # 设置富文本
                            if rich_text_parts:
                                cell.value = CellRichText(*rich_text_parts)
                                
                                # 如果有行内代码，设置背景色
                                if has_inline_code:
                                    cell.fill = code_fill
                        elif len(cell_format.segments) == 1:
                            # 单个片段
                            seg = cell_format.segments[0]
                            cell.value = clean_text
                            
                            # 检查是否有行内代码
                            has_inline_code = seg.is_code
                            if has_inline_code:
                                cell.fill = code_fill
                            
                            # 应用整体格式
                            if seg.bold or seg.italic or seg.strikethrough or seg.is_code:
                                cell.font = Font(
                                    bold=seg.bold,
                                    italic=seg.italic,
                                    strike=seg.strikethrough,
                                    name="Consolas" if seg.is_code else None
                                )
                        else:
                            # 没有格式片段，直接设置值
                            cell.value = clean_text
                    else:
                        # 不保留格式，清除 Markdown 符号
                        cell_format = CellFormat(cell_value)
                        cell.value = cell_format.parse()
                    
                    # 第一行应用表头样式
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                    
                    # 默认居中对齐
                    if not cell.alignment or not cell.alignment.wrap_text:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # 自动调整列宽
            for col_idx in range(1, len(table_data[0]) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for row_idx in range(1, len(table_data) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value:
                            # 考虑换行符
                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines) if lines else 0
                            if max_line_length > max_length:
                                max_length = max_line_length
                    except Exception:
                        pass
                
                # 设置列宽（最小10，最大50）
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # 保存文件
            wb.save(output_path)
            log(f"Successfully generated XLSX with formatting: {output_path}")
            return True
            
        except Exception as e:
            log(f"Failed to generate XLSX: {e}")
            raise InsertError(f"生成 XLSX 文件失败: {e}")

